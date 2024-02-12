import csv
import os
from openpyxl import load_workbook
from pypdf import PdfReader


def test_xlsx_file(create_archive):
    path = os.path.join(create_archive, 'file_example_XLSX_50.xlsx')
    open_xlsx = load_workbook(path)
    sheet = open_xlsx.active
    name = sheet.cell(row=18, column=3).value
    assert name == 'Benito'


def test_users_csv(create_archive):
    path = os.path.join(create_archive, 'csv_file.csv')
    with open(path) as csvfile:
        reader = csv.DictReader(csvfile, delimiter=';')
        names = [row['Username'] for row in reader]
    assert names == ['broker1']


def test_pdf_page_content(create_archive):
    path = f'{create_archive}/Python Testing with Pytest (Brian Okken).pdf'
    text_to_search = "About the Pragmatic Bookshelf"
    with open(path, 'rb') as file:
        reader = PdfReader(file)
        text_found = any(text_to_search in page.extract_text() for page in reader.pages)
        assert text_found, f"Текст \"{text_to_search}\" не найден в PDF файле."
